import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
import util.funcoes as f
from loguru import logger as Logs
from win10toast import ToastNotifier
from datetime import datetime

toaster = ToastNotifier()

vNomeAutomacao = 'RPA_SolicitacaoTitulos'
pathLocalLog = f"C:/Temp/RPA_SolicitacaoTitulos/"
f.pastaExiste(pathLocalLog, True)
nameLog = f"{pathLocalLog}{datetime.today().strftime('%d_%m_%Y')}.log"
Logs.add(nameLog, level="INFO")

sg.theme('Reddit')   # Add a touch of color
# All the stuff inside your window.
# event, values = sg.Window('Get filename example', [[sg.Text('Filename')], , [sg.OK(), sg.Cancel()] ]).read(close=True)
layout = [
            [sg.Text('E-mail da Cooperativa', size=20), sg.Input(key='cpoEmail')],
            [sg.Text('Arquivo de Títulos em Excel', size=20), sg.Input(key='cpoCaminhoArquivo'), sg.FileBrowse('Buscar', file_types=['xls, xlsx'])],
            [sg.Button('Enviar E-mail', size=10, key='btnEnviarEmail'), sg.Button('Cancelar', size=10, key='btnCancelar')] ]

# Create the Window
window = sg.Window('RPA - Solicitação de Títulos', layout)


# Event Loop to process "events" and get the "values" of the inputs
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'btnCancelar': # if user closes window or clicks cancel
        break
    else:

        if len(str(values['cpoEmail']).strip()) == 0:
            sg.popup('Informe um campo de E-mail')
            continue

        if len(str(values['cpoCaminhoArquivo']).strip()) == 0:
            sg.popup('Informe o Local do Arquivo')
            continue

        # Abrir Excel
        vexcel = load_workbook(values['cpoCaminhoArquivo'], data_only=True)
        vplanilha = vexcel.active
        vBobyMail = ''
        for row in range(vplanilha.max_row):
            for column in "A":
                cell_name = "{}{}".format(column, row + 1)
                vBobyMail = vBobyMail + vplanilha[cell_name].value + ';'
        # Montar Corpo do E-mail
        # Enviar E-mail
        vAssunto = f"FICHA GRAFICA AUTOMATIZADA"
        vMensagem = vBobyMail[:-1]
        try:
            f.envia_email(vusuario="lffruhling@hotmail.com", vsenha="pos28680lncymftw", vassunto=vAssunto,
                          vdestinatarios=values['cpoEmail'], vmensagem=vMensagem, vQuebraLinha=False)

            toaster.show_toast("RPA - Solicitação de Títulos",
                               f'O processamento terminou! Um e-mail com os títulos foi enviado para {values["cpoEmail"]}',
                               duration=10)
            break

        except Exception as e:
            toaster.show_toast("RPA - Solicitação de Títulos",
                               f'Falha ao enviar e-mail de solicitação de títulos! Entre em contato com o Suporte',
                               duration=10)
            Logs.error(f'Falha ao enviar e-mail para: {values["cpoEmail"]}')
            Logs.critical(e)
            break

window.close()